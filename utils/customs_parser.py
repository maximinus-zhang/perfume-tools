# -*- coding: utf-8 -*-
"""
customs_parser.py — 海口海关《离岛免税销售情况表》XLSX 解析 + 在线抓取
=================================================================
解析本地已下载的海关月报 XLSX，归一化金额单位（2026=亿元 / 2025=万元），
输出结构化月度记录，供 Streamlit 页面「刷新」直接使用。
同时提供 try_cdp_fetch()：当本机开启调试端口 Chrome 时，在线拉取最新月报。

XLSX 结构（实测）：
  标题行(row1): "2026年4月海南离岛免税销售情况表"  → 解析 年/月
  row4-6 col2 : 免税购物金额 / 免税购物实际人次 / 免税购物件数
  col4 = 本月值, col5 = 本月同比±% , col6 = 累计, col7 = 累计同比±%
  单位(row3): 金额 2026="亿元" / 2025="万元"(须÷10000)；人次="万人次"；件数="万件"
"""
import os
import re
import json
import time
import urllib.request

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import websocket
    HAS_WS = True
except ImportError:
    HAS_WS = False

_HERE = os.path.dirname(os.path.abspath(__file__))
XLSX_DIR = os.path.abspath(os.path.join(_HERE, "..", "data", "hainan_dutyfree", "xlsx"))

CDP = "http://127.0.0.1:9222"
LIST_URL = "http://haikou.customs.gov.cn/haikou_customs/605737/fdzdgknr82/605745/index.html"
DETAIL_URL = "http://haikou.customs.gov.cn/haikou_customs/605737/fdzdgknr82/605745/{0}/index.html"

# (标签, 详情页ID) — 已知月份；新增请用 try_cdp_fetch(discover=True) 自动发现
PAGES = [
    ("2026M01", "7037278"),
    ("2026M02", "7074681"),
    ("2026M03", "7117655"),
    ("2026M04", "7158589"),
    ("2026M05", "7217545"),
    ("2025M11", "6896365"),
    ("2025M12", "6952249"),
]


# ============================================================
# 📖 解析
# ============================================================

def _norm_amount(value, unit):
    """金额单位归一化为亿元。"""
    if value is None:
        return None
    try:
        v = float(value)
    except (TypeError, ValueError):
        return None
    if unit and "万元" in unit:
        return round(v / 10000.0, 4)
    return round(v, 4)


def parse_one_xlsx(path):
    if not HAS_OPENPYXL:
        return None
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    title = ws.cell(row=1, column=1).value or ""
    ym = re.search(r"(\d{4})年(\d{1,2})月", str(title))
    year = int(ym.group(1)) if ym else None
    month = int(ym.group(2)) if ym else None
    rec = {
        "file": os.path.basename(path),
        "year": year,
        "month": month,
        "amt": None, "amt_yoy": None,
        "pax": None, "pax_yoy": None,
        "pieces": None, "pieces_yoy": None,
        "amt_unit": None,
        "ytd_amt": None, "ytd_amt_yoy": None,  # 累计列（如有）
    }
    for r in range(4, min(ws.max_row, 9) + 1):
        label = str(ws.cell(row=r, column=2).value or "")
        unit = str(ws.cell(row=r, column=3).value or "")
        v_cur = ws.cell(row=r, column=4).value
        v_yoy = ws.cell(row=r, column=5).value
        v_ytd = ws.cell(row=r, column=6).value if ws.max_column >= 6 else None
        v_ytd_yoy = ws.cell(row=r, column=7).value if ws.max_column >= 7 else None
        if "金额" in label:
            rec["amt"] = _norm_amount(v_cur, unit)
            rec["amt_yoy"] = v_yoy
            rec["amt_unit"] = "亿元"  # 已统一归一化为亿元
            rec["ytd_amt"] = _norm_amount(v_ytd, unit)
            rec["ytd_amt_yoy"] = v_ytd_yoy
        elif "人次" in label:
            rec["pax"] = v_cur
            rec["pax_yoy"] = v_yoy
        elif "件数" in label:
            rec["pieces"] = v_cur
            rec["pieces_yoy"] = v_yoy
    return rec


def parse_customs_folder(folder=None):
    folder = os.path.abspath(folder or XLSX_DIR)
    if not os.path.isdir(folder):
        return []
    recs = []
    for fn in sorted(os.listdir(folder)):
        if fn.lower().endswith(".xlsx") and not fn.startswith("~"):
            rec = parse_one_xlsx(os.path.join(folder, fn))
            if rec and rec["year"]:
                recs.append(rec)
    return recs


def build_monthly_from_xlsx(folder=None):
    """返回 (by_ym{(year,month):rec}, recs_list)。"""
    recs = parse_customs_folder(folder)
    by_ym = {(r["year"], r["month"]): r for r in recs}
    return by_ym, recs


# ============================================================
# 🌐 在线抓取（CDP，需本机开调试 Chrome）
# ============================================================

def cdp_available():
    try:
        with urllib.request.urlopen(f"{CDP}/json", timeout=3) as r:
            return r.status == 200
    except Exception:
        return False


def _download(url, outp):
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=30) as resp:
        data = resp.read()
    with open(outp, "wb") as f:
        f.write(data)
    return len(data)


def _cdp_download_one(ws, send, wait, label, pid, out_dir):
    url = DETAIL_URL.format(pid)
    # 先导航 + 等待加载
    wid = send("Page.navigate", {"url": url})
    wait(wid, 10)
    time.sleep(4)
    expr = ("(()=>{var a=[].slice.call(document.querySelectorAll('a')).map(function(x){return x.href;})"
            ".filter(function(h){return h&&h.toLowerCase().indexOf('.xlsx')>-1;});return JSON.stringify(a);})()")
    wid = send("Runtime.evaluate", {"expression": expr, "returnByValue": True})
    res = wait(wid, 8)
    links = []
    if res and "result" in res:
        try:
            links = json.loads(res["result"].get("result", {}).get("value", "[]"))
        except Exception:
            links = []
    if not links:
        return False, "未找到 xlsx 链接"
    xurl = links[0]
    outp = os.path.join(out_dir, f"{label}.xlsx")
    try:
        sz = _download(xurl, outp)
        return (sz > 1000), f"下载 {sz}B"
    except Exception as e:
        return False, f"下载失败: {e}"


def try_cdp_fetch(folder=None, discover=False):
    """
    在线抓取海关月报。返回 (success:bool, message:str)。
    前置：本机开启调试 Chrome（--remote-debugging-port=9222 --remote-allow-origins=*）。
    """
    folder = os.path.abspath(folder or XLSX_DIR)
    os.makedirs(folder, exist_ok=True)
    if not HAS_WS:
        return False, "未安装 websocket-client（pip install websocket-client）"
    if not cdp_available():
        return False, "本机未检测到调试 Chrome（127.0.0.1:9222），无法在线抓取；请用『重新解析本地月报』"

    import websocket as _ws_mod
    pages = list(PAGES)
    if discover:
        try:
            with urllib.request.urlopen(f"{CDP}/json", timeout=5) as r:
                tabs = json.load(r)
            page = next((t for t in tabs if t.get("type") == "page" and "webSocketDebuggerUrl" in t), None)
            if page:
                ws0 = _ws_mod.create_connection(page["webSocketDebuggerUrl"], timeout=30)
                mid = [1]
                def snd(m, p=None):
                    mid[0] += 1
                    ws0.send(json.dumps({"id": mid[0], "method": m, "params": p or {}}))
                    return mid[0]
                def w8(wid, t=15):
                    end = time.time() + t
                    while time.time() < end:
                        try:
                            raw = ws0.recv()
                        except Exception:
                            continue
                        o = json.loads(raw)
                        if o.get("id") == wid:
                            return o
                    return None
                snd("Page.enable"); snd("Runtime.enable")
                expr = ("(()=>{var re=[];var as=[].slice.call(document.querySelectorAll('a'));"
                        "for(var i=0;i<as.length;i++){var h=as[i].href||'';var t=as[i].innerText||'';"
                        "var m=h.match(/605745\\/(\\d+)\\/index\\.html/);"
                        "if(m){re.push(['" + "','" + "'].join([t.trim().replace(/\\s+/g,' '), m[1]]));}}return JSON.stringify(re);})()")
                wid = snd("Runtime.evaluate", {"expression": expr, "returnByValue": True})
                res = w8(wid, 12)
                if res and "result" in res:
                    try:
                        found = json.loads(res["result"].get("result", {}).get("value", "[]"))
                        known = {p for _, p in PAGES}
                        for t, pid in found:
                            if pid not in known:
                                pages.append((t, pid))
                    except Exception:
                        pass
                ws0.close()
        except Exception:
            pass

    try:
        with urllib.request.urlopen(f"{CDP}/json", timeout=5) as r:
            tabs = json.load(r)
        page = next((t for t in tabs if t.get("type") == "page" and "webSocketDebuggerUrl" in t), None)
        if not page:
            return False, "没有可用的 page target"
        ws = _ws_mod.create_connection(page["webSocketDebuggerUrl"], timeout=30)
        mid = [1]
        def send(m, p=None):
            mid[0] += 1
            ws.send(json.dumps({"id": mid[0], "method": m, "params": p or {}}))
            return mid[0]
        def wait(wid, t=20):
            end = time.time() + t
            while time.time() < end:
                try:
                    raw = ws.recv()
                except Exception:
                    continue
                o = json.loads(raw)
                if o.get("id") == wid:
                    return o
            return None
        send("Page.enable"); send("Runtime.enable")
        ok = 0
        msgs = []
        for label, pid in pages:
            succ, msg = _cdp_download_one(ws, send, wait, label, pid, folder)
            if succ:
                ok += 1
            msgs.append(f"{label}:{'✓' if succ else '✗'}{msg}")
        ws.close()
        return True, f"在线抓取完成 {ok}/{len(pages)}：\n" + "\n".join(msgs)
    except Exception as e:
        return False, f"CDP 抓取异常: {e}"


if __name__ == "__main__":
    by_ym, recs = build_monthly_from_xlsx()
    print(f"解析到 {len(recs)} 个月报：")
    for r in recs:
        print(f"  {r['year']}-{r['month']:02d}  金额={r['amt']}({r['amt_unit']}) 同比={r['amt_yoy']}%  "
              f"人次={r['pax']} 件数={r['pieces']}  累计={r['ytd_amt']}")
