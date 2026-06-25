import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# 生成采购数据 Excel 模板
# ============================================================

wb = Workbook()

# ============================================================
# Sheet 1: 使用说明
# ============================================================
ws1 = wb.active
ws1.title = "使用说明"

# 样式定义
title_font = Font(name='微软雅黑', size=14, bold=True, color='FFFFFF')
title_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
header_font = Font(name='微软雅黑', size=11, bold=True, color='1F3864')
header_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
normal_font = Font(name='微软雅黑', size=10)
required_font = Font(name='微软雅黑', size=10, color='C00000')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# 标题行
ws1.merge_cells('A1:F1')
cell = ws1['A1']
cell.value = '📦 香水供应链 · 采购数据模板 - 使用说明'
cell.font = Font(name='微软雅黑', size=16, bold=True, color='2F5496')
cell.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 40

# 说明内容
instructions = [
    ('📋 模板说明', ''),
    ('本模板包含 2 个工作表（Sheet），请按说明填写：', ''),
    ('', ''),
    ('━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━', ''),
    ('【Sheet 1: 使用说明】', '当前页面 — 填写前请仔细阅读'),
    ('', ''),
    ('【Sheet 2: 采购数据（主表）】', '在此表中填写实际采购数据'),
    ('', ''),
    ('━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━', ''),
    ('📌 字段说明（共11列）：', ''),
    ('', ''),
]

# 字段说明表
fields = [
    ('列号', '字段名', '格式/示例', '必填', '说明'),
    ('A', '采购单号', 'PO-202506-0001', '✅', '唯一编号，不可重复'),
    ('B', '品牌', 'CHANEL', '✅', '请统一使用品牌英文名'),
    ('C', '产品类别', '香水EDP', '✅', '香水EDP/香水EDT/古龙水/旅行装/礼盒套装'),
    ('D', '供应商', '法国香奈儿集团', '✅', '填写供应商全称'),
    ('E', '采购数量', '500', '✅', '正整数'),
    ('F', '单价(CNY)', '320.00', '✅', '正数，保留2位小数'),
    ('G', '总金额(CNY)', '160000', '❌', '系统自动 = 数量×单价，可不填'),
    ('H', '下单日期', '2025-06-01', '✅', '格式 YYYY-MM-DD'),
    ('I', '预计到货', '2025-06-15', '✅', '格式 YYYY-MM-DD'),
    ('J', '实际到货', '2025-06-14 或留空', '❌', '未到货则留空'),
    ('K', '状态', '已到货', '✅', '待审核/待发货/运输中/已到货/已取消'),
]

# 写入说明文本
row = 3
for text, _ in instructions:
    ws1.merge_cells(f'A{row}:F{row}')
    cell = ws1[f'A{row}']
    cell.value = text
    if text.startswith('━'):
        cell.font = Font(name='微软雅黑', size=10, color='888888')
    elif text.startswith('📌'):
        cell.font = Font(name='微软雅黑', size=12, bold=True, color='2F5496')
    else:
        cell.font = normal_font
    row += 1

# 写入字段说明表头
row += 1
for col_idx, header in enumerate(fields[0], 1):
    cell = ws1.cell(row=row, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')

# 写入字段说明数据
for data in fields[1:]:
    row += 1
    for col_idx, value in enumerate(data, 1):
        cell = ws1.cell(row=row, column=col_idx, value=value)
        cell.font = required_font if value == '✅' else normal_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

# 调整列宽
ws1.column_dimensions['A'].width = 10
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 25
ws1.column_dimensions['D'].width = 10
ws1.column_dimensions['E'].width = 40
ws1.column_dimensions['F'].width = 40

# ============================================================
# Sheet 2: 采购数据（主表）
# ============================================================
ws2 = wb.create_sheet(title="采购数据")

# 列名
headers = ['采购单号', '品牌', '产品类别', '供应商', '采购数量', 
           '单价(CNY)', '总金额(CNY)', '下单日期', '预计到货', '实际到货', '状态']

# 写入列名
for col_idx, header in enumerate(headers, 1):
    cell = ws2.cell(row=1, column=col_idx, value=header)
    cell.font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')

# 示例数据
sample_data = [
    ['PO-202506-0001', 'CHANEL', '香水EDP', '法国香奈儿集团', 500, 320, 160000, '2025-06-01', '2025-06-15', '2025-06-14', '已到货'],
    ['PO-202506-0002', 'DIOR', '香水EDT', 'LVMH集团', 300, 280, 84000, '2025-06-03', '2025-06-18', '', '运输中'],
    ['PO-202506-0003', 'HERMES', '古龙水', '爱马仕国际', 200, 450, 90000, '2025-06-05', '2025-06-20', '', '待发货'],
    ['PO-202506-0004', 'JO MALONE', '礼盒套装', '雅诗兰黛集团', 400, 380, 152000, '2025-06-08', '2025-06-22', '', '待审核'],
    ['', '', '', '', '', '', '', '', '', '', ''],
    ['', '', '', '', '', '', '', '', '', '', ''],
    ['', '', '', '', '', '', '', '', '', '', ''],
    ['📝 填写说明：', '', '', '', '', '', '', '', '', '', ''],
    ['• 请从第2行开始填写，不要修改第1行列名', '', '', '', '', '', '', '', '', '', ''],
    ['• 采购单号请保持唯一，不要重复', '', '', '', '', '', '', '', '', '', ''],
    ['• 总金额(CNY)可不填，上传后系统会自动计算', '', '', '', '', '', '', '', '', '', ''],
    ['• 日期格式固定为 YYYY-MM-DD（如 2025-06-01）', '', '', '', '', '', '', '', '', '', ''],
    ['• 实际到货留空表示尚未到货', '', '', '', '', '', '', '', '', '', ''],
]

# 写入示例数据
for row_idx, data in enumerate(sample_data, 2):
    for col_idx, value in enumerate(data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        if row_idx <= 5:  # 示例数据
            cell.font = Font(name='微软雅黑', size=10, color='666666')
            cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        else:  # 说明文本
            cell.font = Font(name='微软雅黑', size=10, color='2F5496', italic=True)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

# 调整列宽
col_widths = [18, 15, 12, 20, 12, 12, 15, 14, 14, 14, 12]
for i, width in enumerate(col_widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = width

# ============================================================
# 保存文件
# ============================================================
output_file = 'data/purchase_data.xlsx'
wb.save(output_file)
print(f"✅ 模板已生成：{output_file}")
print(f"   Sheet 1: 使用说明（字段说明、填写注意事项）")
print(f"   Sheet 2: 采购数据（示例数据 + 空行供填写）")
