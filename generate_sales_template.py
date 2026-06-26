import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import random

wb = Workbook()

# ============================================================
# Sheet 1: 使用说明
# ============================================================
ws1 = wb.active
ws1.title = "使用说明"

title_font = Font(name='微软雅黑', size=14, bold=True, color='FFFFFF')
header_font = Font(name='微软雅黑', size=11, bold=True, color='1F3864')
header_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
normal_font = Font(name='微软雅黑', size=10)
required_font = Font(name='微软雅黑', size=10, color='C00000')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

ws1.merge_cells('A1:E1')
cell = ws1['A1']
cell.value = '📈 香水供应链 · 历史销售数据模板 - 使用说明'
cell.font = Font(name='微软雅黑', size=16, bold=True, color='2F5496')
cell.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 40

row = 3
for text in [
    '📋 模板说明',
    '本模板包含 2 个工作表（Sheet），请按说明填写：',
    '',
    '━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━',
    '【Sheet 1: 使用说明】当前页面 — 填写前请仔细阅读',
    '',
    '【Sheet 2: 销售数据】在此表中填写历史销售数据',
    '',
    '━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━',
    '📌 字段说明（共3列）：',
    '',
]:
    ws1.merge_cells(f'A{row}:E{row}')
    cell = ws1[f'A{row}']
    cell.value = text
    if text.startswith('━'):
        cell.font = Font(name='微软雅黑', size=10, color='888888')
    elif text.startswith('📌'):
        cell.font = Font(name='微软雅黑', size=12, bold=True, color='2F5496')
    else:
        cell.font = normal_font
    row += 1

fields = [
    ('列号', '字段名', '格式/示例', '必填', '说明'),
    ('A', '日期', '2025-01-01', '✅', '格式 YYYY-MM-DD，每天一行'),
    ('B', '销量', '150', '✅', '正整数，该日销售数量'),
    ('C', '备注（可选）', '春节促销', '❌', '可选，标记特殊日期'),
]

row += 1
for col_idx, header in enumerate(fields[0], 1):
    cell = ws1.cell(row=row, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')

for data in fields[1:]:
    row += 1
    for col_idx, value in enumerate(data, 1):
        cell = ws1.cell(row=row, column=col_idx, value=value)
        cell.font = required_font if value == '✅' else normal_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

ws1.column_dimensions['A'].width = 10
ws1.column_dimensions['B'].width = 15
ws1.column_dimensions['C'].width = 30
ws1.column_dimensions['D'].width = 10
ws1.column_dimensions['E'].width = 40

# ============================================================
# Sheet 2: 销售数据
# ============================================================
ws2 = wb.create_sheet(title="销售数据")

headers = ['日期', '销量', '备注']
for col_idx, header in enumerate(headers, 1):
    cell = ws2.cell(row=1, column=col_idx, value=header)
    cell.font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')

# 生成示例数据（2025年1月到2026年6月，约540天）
start_date = datetime(2025, 1, 1)
sample_data = []
for i in range(540):
    d = start_date + timedelta(days=i)
    # 模拟有趋势的销售数据
    base = 50 + i * 0.05  # 缓慢上升趋势
    season = 15 * (2 ** 0.5 * ((i % 365) / 365 * 2 - 1))  # 季节性波动
    noise = random.gauss(0, 8)
    sales = max(5, int(base + season + noise))
    
    # 节假日效应
    remark = ''
    month_day = (d.month, d.day)
    if month_day == (1, 1):
        sales = int(sales * 1.3)
        remark = '元旦'
    elif month_day == (2, 10):
        sales = int(sales * 1.5)
        remark = '春节'
    elif month_day == (10, 1):
        sales = int(sales * 1.4)
        remark = '国庆'
    elif month_day == (5, 1):
        sales = int(sales * 1.2)
        remark = '劳动节'
    elif month_day == (6, 18):
        sales = int(sales * 1.3)
        remark = '618促销'
    elif month_day == (11, 11):
        sales = int(sales * 1.6)
        remark = '双11'
    
    sample_data.append([d.strftime('%Y-%m-%d'), sales, remark])

for row_idx, data in enumerate(sample_data, 2):
    for col_idx, value in enumerate(data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.font = Font(name='微软雅黑', size=10, color='666666')
        cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

# 留空行
for row_idx in range(542, 550):
    for col_idx in range(1, 4):
        cell = ws2.cell(row=row_idx, column=col_idx, value='')
        cell.border = thin_border

# 填写提示
tips_row = 552
tips = [
    '📝 填写说明：',
    '• 请从第2行开始填写，不要修改第1行列名',
    '• 日期格式固定为 YYYY-MM-DD（如 2025-01-01）',
    '• 销量为当天销售数量，填写正整数',
    '• 备注列可选，可标记促销活动、节假日等特殊事件',
    '• 数据越完整，预测结果越准确（建议至少半年以上数据）',
]
for i, tip in enumerate(tips, tips_row):
    ws2.merge_cells(f'A{i}:C{i}')
    cell = ws2.cell(row=i, column=1, value=tip)
    cell.font = Font(name='微软雅黑', size=10, color='2F5496', italic=True)
    cell.alignment = Alignment(horizontal='left', vertical='center')

ws2.column_dimensions['A'].width = 15
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 20

# 保存
output_file = 'data/sales_history.xlsx'
wb.save(output_file)
print(f"✅ 销售历史模板已生成：{output_file}")
print(f"   包含 {len(sample_data)} 行示例数据（2025-01-01 ~ 2026-06-23）")
print(f"   Sheet 1: 使用说明")
print(f"   Sheet 2: 销售数据（日期、销量、备注）")
