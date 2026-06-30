"""
📊 订单满足率分析工具 - Streamlit 页面
========================================
功能：上传订单+库存文件，计算TR总库存满足量 + 可借货仓满足量
"""
import streamlit as st
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
import re
import io
import zipfile
import tempfile
import shutil
import time
import traceback

# ===================== 配置 =====================
st.set_page_config(page_title="订单满足率分析", page_icon="✅", layout="wide")

BORROWABLE_WH = [
    'W-EC-001', 'W-EC-002',
    'W-FG-001', 'W-FG-002', 'W-FG-004',
    'X-FG-002',
    'Y-NL-001', 'Y-NL-002', 'Y-NL-004', 'Y-NL-005', 'Y-NL-008'
]

# ===================== 工具函数 =====================

def clean_code(code):
    if not code: return code
    s = str(code).strip()
    s = re.sub(r'-[A-Za-z]$', '', s)
    s = re.sub(r'-[a-z]$', '', s)
    if s.isdigit() and s.startswith('0'):
        s = s.lstrip('0')
        if s == '': s = '0'
    return s


def safe_val(ws, r, c):
    try:
        return ws.cell(row=r, column=c).value
    except AttributeError:
        return None


def load_inventory(uploaded_file):
    """从上传的库存文件加载⑤指定仓库分布sheet"""
    df = pd.read_excel(uploaded_file, sheet_name="⑤指定仓库分布", engine='openpyxl')
    cols = list(df.columns)
    
    # 构建仓库代码 → 实际列名映射
    wh_col_map = {}
    for col in cols:
        col_str = str(col)
        for wh_code in BORROWABLE_WH:
            if col_str.startswith(wh_code):
                wh_col_map[wh_code] = col_str
                break
    
    # 找到 TR总库存 列
    tr_col = None
    for c in cols:
        if 'TR' in str(c) and ('总' in str(c) or '库存' in str(c)):
            tr_col = c
            break
    if tr_col is None and 'TR总库存' in cols:
        tr_col = 'TR总库存'
    
    # 构建索引
    inventory = {}
    for _, row in df.iterrows():
        code = str(row.get('orig_code', '') or row.get('sap_code', '') or '').strip()
        if not code: continue
        
        cleaned = clean_code(code)
        if cleaned not in inventory:
            inventory[cleaned] = {wh: 0 for wh in BORROWABLE_WH}
            if tr_col:
                inventory[cleaned]['TR总库存'] = 0
            inventory[cleaned]['_orig_codes'] = set()
        
        inventory[cleaned]['_orig_codes'].add(code)
        
        if tr_col:
            try:
                inventory[cleaned]['TR总库存'] += float(row.get(tr_col, 0) or 0)
            except: pass
        
        for wh in BORROWABLE_WH:
            actual_col = wh_col_map.get(wh, wh)
            try:
                v = float(row.get(actual_col, 0) or 0)
                if v > 0:
                    inventory[cleaned][wh] += v
            except: pass
    
    return inventory, tr_col, len(df)


def calculate_fulfillment(order_qty, inv_record, tr_col):
    if inv_record is None:
        return 0, "无库存", order_qty, "", 0, "库存表中无此SKU"
    
    if order_qty <= 0:
        return 0, "", 0, "", 0, ""
    
    remaining = order_qty
    fulfillment_detail = []
    
    # 1. 优先用 TR总库存
    tr_qty = int(inv_record.get('TR总库存', 0)) if tr_col else 0
    if tr_qty > 0:
        take = min(remaining, tr_qty)
        if take > 0:
            fulfillment_detail.append((take, f"TR总库存"))
            remaining -= take
    
    if remaining <= 0:
        return order_qty, f"{order_qty}@TR总库存", 0, "", 0, ""
    
    # 2. 从可借货仓取
    borrowed_wh = None
    borrowed_wh_stock = 0
    
    wh_stocks = [(wh, int(inv_record.get(wh, 0))) for wh in BORROWABLE_WH]
    wh_stocks = [(wh, q) for wh, q in wh_stocks if q > 0]
    wh_stocks.sort(key=lambda x: x[1], reverse=True)
    
    if wh_stocks:
        best_wh, best_qty = wh_stocks[0]
        borrowed_wh_stock = best_qty
        take = min(remaining, best_qty)
        borrowed_wh = best_wh
        fulfillment_detail.append((take, f"借货-{best_wh}"))
        remaining -= take
    
    fulfilled = order_qty - remaining
    source_str = " + ".join([f"{q}@{w}" for q, w in fulfillment_detail])
    
    remark_parts = []
    if wh_stocks and borrowed_wh:
        for wh, qty in wh_stocks:
            if wh != borrowed_wh:
                remark_parts.append(f"{wh}(可借{qty})")
    if remaining > 0:
        remark_parts.append(f"仍缺{remaining}")
    remark = "; ".join(remark_parts)
    
    return fulfilled, source_str, remaining, borrowed_wh or "", borrowed_wh_stock, remark


# ===================== 订单解析函数 =====================

def find_header_row(ws, max_scan=10):
    for r in range(1, min(max_scan + 1, ws.max_row + 1)):
        for c in range(1, min(ws.max_column + 1, 25)):
            v = safe_val(ws, r, c)
            if v:
                s = str(v).strip().lower()
                if "reference" in s or "产品编号" in s:
                    return r
    return None


def find_col_by_keyword(ws, header_row, keywords, max_scan=40):
    for c in range(1, min(max_scan + 1, ws.max_column + 1)):
        v = safe_val(ws, header_row, c)
        if v:
            s = str(v).strip().lower()
            for kw in keywords:
                if s == kw.lower() or kw.lower() in s:
                    return c
    return None


def get_sheet_order_data(ws, sn):
    if ws.max_row < 5 or ws.max_column < 3:
        return None
    
    sn_lower = sn.lower().strip()
    skip_names = ["summary", "gwp", "满足率", "code check", "order summary", "汇总"]
    if any(skip in sn_lower for skip in skip_names):
        return None
    
    # SAP PO 格式
    r1_vals = {}
    for c in range(1, min(ws.max_column + 1, 50)):
        v = safe_val(ws, 1, c)
        if v:
            r1_vals[str(v).strip().lower()] = c
    
    if "po quantity" in r1_vals:
        cc = r1_vals.get("vendor article no") or r1_vals.get("vendor article no.") or r1_vals.get("article")
        qc = r1_vals.get("po quantity")
        if cc and qc:
            data = []
            for r in range(2, ws.max_row + 1):
                cv = safe_val(ws, r, cc)
                qv = safe_val(ws, r, qc)
                if cv is None: continue
                cs = str(cv).strip()
                if not cs or len(cs) < 3: continue
                try:
                    q = float(str(qv).strip()) if qv else 0
                    if q > 0: data.append((r, cs, int(q)))
                except: pass
            if data: return data
    
    # DFS 格式
    r1_keys = {str(safe_val(ws, 1, c) or '').strip().lower(): c for c in range(1, min(ws.max_column + 1, 50))}
    
    if "department" in r1_keys and "class" in r1_keys:
        for r in range(2, min(ws.max_row + 1, 100)):
            dv = str(safe_val(ws, r, r1_keys.get("department", 0)) or '').strip().lower()
            if dv == "total":
                max_data_row = r - 1
                break
        else:
            max_data_row = ws.max_row
        
        d_col = r1_keys.get("style code") or r1_keys.get("style no") or r1_keys.get("product code") or None
        q_col = r1_keys.get("order qty") or r1_keys.get("qty") or r1_keys.get("quantity") or None
        if d_col and q_col:
            data = []
            for r in range(2, max_data_row + 1):
                cv = safe_val(ws, r, d_col)
                qv = safe_val(ws, r, q_col)
                if cv is None or qv is None: continue
                cs = str(cv).strip()
                if not cs or len(cs) < 3: continue
                try:
                    q = float(str(qv).strip())
                    if q > 0: data.append((r, cs, int(q)))
                except: pass
            if data: return data
    
    # CDFG 标准格式
    header_row = find_header_row(ws)
    if header_row:
        ref_col = find_col_by_keyword(ws, header_row, ["reference", "ref #"])
        if ref_col:
            qty_col = find_col_by_keyword(ws, header_row, ["qty"])
            if qty_col:
                data = []
                for r in range(header_row + 1, ws.max_row + 1):
                    cv = safe_val(ws, r, ref_col)
                    qv = safe_val(ws, r, qty_col)
                    if cv is None: continue
                    cs = str(cv).strip()
                    if not cs or cs.lower() in ['', 'total', 'subtotal', '合计']: continue
                    try:
                        q = float(str(qv or '0').strip())
                        if q > 0: data.append((r, cs, int(q)))
                    except: pass
                if data: return data
    
    # GDF 格式
    for r in range(1, min(6, ws.max_row + 1)):
        for c in range(1, min(ws.max_column + 1, 30)):
            v = safe_val(ws, r, c)
            if v and str(v).strip().upper() == "VAN":
                header_row = r
                van_col = c
                qty_candidates = []
                for c2 in range(c + 1, min(ws.max_column + 1, c + 10)):
                    v2 = safe_val(ws, r, c2)
                    if v2 and "qty" in str(v2).strip().lower():
                        qty_candidates.append(c2)
                if not qty_candidates:
                    for c2 in range(c + 1, min(ws.max_column + 1, c + 10)):
                        v2 = safe_val(ws, r, c2)
                        if v2 and "quantity" in str(v2).strip().lower():
                            qty_candidates.append(c2)
                if qty_candidates:
                    qty_col = qty_candidates[0]
                    data = []
                    for r2 in range(header_row + 1, ws.max_row + 1):
                        cv = safe_val(ws, r2, van_col)
                        qv = safe_val(ws, r2, qty_col)
                        if cv is None: continue
                        cs = str(cv).strip()
                        if not cs or len(cs) < 3: continue
                        try:
                            q = float(str(qv or '0').strip())
                            if q > 0: data.append((r2, cs, int(q)))
                        except: pass
                    if data: return data
    
    return None


def process_order_file(filepath, inventory, tr_col):
    """处理单个订单文件"""
    wb = load_workbook(filepath, data_only=True)
    sheets_processed = 0
    total_skus = 0
    total_qty = 0
    total_fulfilled = 0
    summary_lines = []
    
    for sn in wb.sheetnames:
        ws = wb[sn]
        
        order_data = get_sheet_order_data(ws, sn)
        if order_data is None:
            continue
        
        last_col = ws.max_column
        header_added = False
        match_count = 0
        sheet_qty = 0
        sheet_fulfilled = 0
        
        for row_num, code, qty in order_data:
            cleaned = clean_code(code)
            inv_record = inventory.get(cleaned, None)
            
            fulfilled, source_str, remaining, borrow_wh, borrow_wh_stock, remark = calculate_fulfillment(
                qty, inv_record, tr_col
            )
            
            if not header_added:
                ws.cell(row=1, column=last_col + 1, value="匹配产品CODE")
                ws.cell(row=1, column=last_col + 2, value="满足数量")
                ws.cell(row=1, column=last_col + 3, value="来源仓别")
                ws.cell(row=1, column=last_col + 4, value="需借货数量")
                ws.cell(row=1, column=last_col + 5, value="借货仓别")
                ws.cell(row=1, column=last_col + 6, value="借货仓库存")
                ws.cell(row=1, column=last_col + 7, value="REMARK")
                header_added = True
            
            matched_code = ""
            if inv_record and '_orig_codes' in inv_record:
                orig_set = inv_record['_orig_codes']
                if orig_set:
                    matched_code = next(iter(orig_set))
            
            ws.cell(row=row_num, column=last_col + 1, value=matched_code)
            ws.cell(row=row_num, column=last_col + 2, value=fulfilled)
            ws.cell(row=row_num, column=last_col + 3, value=source_str)
            ws.cell(row=row_num, column=last_col + 4, value=remaining)
            
            if remaining <= 0:
                ws.cell(row=row_num, column=last_col + 5, value="")
                ws.cell(row=row_num, column=last_col + 6, value="")
                ws.cell(row=row_num, column=last_col + 7, value="")
            else:
                ws.cell(row=row_num, column=last_col + 5, value=borrow_wh)
                ws.cell(row=row_num, column=last_col + 6, value=borrow_wh_stock)
                ws.cell(row=row_num, column=last_col + 7, value=remark)
            
            match_count += 1
            sheet_qty += qty
            sheet_fulfilled += fulfilled
        
        if match_count > 0:
            rate = (sheet_fulfilled / sheet_qty * 100) if sheet_qty > 0 else 0
            summary_lines.append(f"Sheet「{sn[:20]}」: {match_count} SKU | {sheet_fulfilled}/{sheet_qty} = {rate:.0f}%")
            sheets_processed += 1
            total_skus += match_count
            total_qty += sheet_qty
            total_fulfilled += sheet_fulfilled
    
    if sheets_processed > 0:
        output_path = str(filepath).replace('.xlsx', '_满足率.xlsx').replace('.XLSX', '_满足率.xlsx')
        wb.save(output_path)
        wb.close()
        return output_path, summary_lines, sheets_processed, total_skus, total_qty, total_fulfilled
    else:
        wb.close()
        return None, ["⚠️ 所有 Sheet 均无订单数据"], 0, 0, 0, 0


# ===================== Streamlit 页面 =====================

st.title("📊 订单满足率分析")
st.markdown("从库存表匹配订单，计算 TR 总库存 + 可借货仓满足情况")

st.divider()

# 创建两列布局
col1, col2 = st.columns(2)

with col1:
    st.subheader("📁 上传订单文件")
    st.caption("支持 Excel (.xlsx) 或 ZIP 压缩包")
    order_files = st.file_uploader(
        "选择文件",
        type=['xlsx', 'zip'],
        accept_multiple_files=True,
        key="order_uploader",
        label_visibility="collapsed"
    )

with col2:
    st.subheader("📦 上传库存表")
    st.caption("需包含「⑤指定仓库分布」sheet")
    inventory_file = st.file_uploader(
        "选择文件",
        type=['xlsx'],
        key="inventory_uploader",
        label_visibility="collapsed"
    )

st.divider()

# 执行按钮
col_left, col_mid, col_right = st.columns([1, 2, 1])
with col_mid:
    execute_btn = st.button("🚀 执行满足率分析", type="primary", use_container_width=True)

# 结果区域
status_box = st.empty()
detail_box = st.empty()
download_box = st.empty()

if execute_btn:
    if not order_files or not inventory_file:
        st.error("⚠️ 请同时上传订单文件和库存表！")
    else:
        try:
            # 1. 加载库存
            with st.spinner("正在加载库存数据..."):
                inventory, tr_col, row_count = load_inventory(inventory_file)
            
            status_box.success(f"✅ 库存加载完成：共 {row_count} 行数据，{len(inventory)} 个唯一 SKU")
            
            # 创建临时目录
            with tempfile.TemporaryDirectory() as tmpdir:
                tmp_dir = Path(tmpdir)
                input_dir = tmp_dir / "input"
                output_dir = tmp_dir / "output"
                input_dir.mkdir()
                output_dir.mkdir()
                
                # 2. 解压/复制订单文件
                file_count = 0
                for uploaded_file in order_files:
                    if uploaded_file.name.endswith('.zip'):
                        with zipfile.ZipFile(uploaded_file, 'r') as zf:
                            zf.extractall(input_dir)
                    else:
                        file_path = input_dir / uploaded_file.name
                        with open(file_path, 'wb') as f:
                            f.write(uploaded_file.getbuffer())
                        file_count += 1
                
                # 收集所有Excel文件
                excel_files = list(input_dir.rglob("*.xlsx")) + list(input_dir.rglob("*.XLSX"))
                if not excel_files:
                    excel_files = list(input_dir.rglob("*.xls")) + list(input_dir.rglob("*.XLS"))
                
                if not excel_files:
                    st.error("❌ 未找到任何 Excel 文件！")
                else:
                    st.info(f"📊 发现 {len(excel_files)} 个 Excel 文件待处理")
                    
                    # 3. 处理每个文件
                    progress_bar = st.progress(0, text="准备处理...")
                    total_files = len(excel_files)
                    
                    all_results = []
                    success_count = 0
                    fail_count = 0
                    grand_total_skus = 0
                    grand_total_qty = 0
                    grand_total_fulfilled = 0
                    
                    result_details = []
                    
                    for idx, fp in enumerate(excel_files):
                        rel_path = fp.relative_to(input_dir)
                        out_path = output_dir / rel_path.parent
                        out_path.mkdir(parents=True, exist_ok=True)
                        
                        # 复制原文件到输出目录
                        shutil.copy2(fp, out_path / fp.name)
                        target_fp = out_path / fp.name
                        
                        progress_bar.progress(
                            (idx) / total_files,
                            text=f"⏳ 处理中 ({idx+1}/{total_files}): {fp.name}"
                        )
                        
                        try:
                            result_path, summary_lines, sheets, skus, qt, ful = process_order_file(
                                str(target_fp), inventory, tr_col
                            )
                            
                            if result_path:
                                all_results.append(Path(result_path))
                                success_count += 1
                                grand_total_skus += skus
                                grand_total_qty += qt
                                grand_total_fulfilled += ful
                                result_details.append((fp.name, sheets, skus, qt, ful))
                                
                                for line in summary_lines:
                                    detail_box.info(f"**{fp.name}** → {line}")
                            else:
                                fail_count += 1
                                for line in summary_lines:
                                    detail_box.warning(f"{fp.name}: {line}")
                                
                        except Exception as e:
                            fail_count += 1
                            detail_box.error(f"❌ {fp.name}: {str(e)}")
                    
                    progress_bar.progress(1.0, text="✅ 处理完成！")
                    
                    # 4. 打包结果
                    if all_results:
                        # 汇总统计
                        overall_rate = (grand_total_fulfilled / grand_total_qty * 100) if grand_total_qty > 0 else 0
                        
                        col_a, col_b, col_c, col_d = st.columns(4)
                        col_a.metric("✅ 成功", f"{success_count} 个文件")
                        col_b.metric("❌ 失败", f"{fail_count} 个文件")
                        col_c.metric("📦 SKU 数", f"{grand_total_skus}")
                        col_d.metric("🎯 总满足率", f"{grand_total_fulfilled}/{grand_total_qty} = {overall_rate:.0f}%")
                        
                        # 显示每个文件的详情
                        with st.expander("📋 各文件处理详情", expanded=True):
                            for fname, sheets, skus, qt, ful in result_details:
                                rate = (ful / qt * 100) if qt > 0 else 0
                                st.write(f"**{fname}** — {sheets} sheet, {skus} SKU | 满足率: {ful}/{qt} = {rate:.0f}%")
                        
                        # 创建ZIP下载包
                        zip_buffer = io.BytesIO()
                        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
                            for result_file in all_results:
                                rel_path = result_file.relative_to(output_dir)
                                zf.write(result_file, str(rel_path))
                        
                        zip_buffer.seek(0)
                        
                        status_box.success(f"✅ 处理完成！成功: {success_count}, 失败: {fail_count}")
                        
                        download_box.download_button(
                            label="📥 下载所有满足率结果 (ZIP)",
                            data=zip_buffer,
                            file_name=f"订单满足率分析_{time.strftime('%Y%m%d_%H%M%S')}.zip",
                            mime="application/zip",
                            use_container_width=True
                        )
                    else:
                        status_box.error("❌ 所有文件处理失败，无结果可下载")
                        
        except Exception as e:
            st.error(f"❌ 发生错误: {str(e)}")
            with st.expander("🔍 错误详情"):
                st.code(traceback.format_exc())

# 底部说明
st.divider()
with st.expander("📖 使用说明"):
    st.markdown("""
    ### 操作步骤
    1. **上传订单文件** — 客户订单 Excel，或包含多个订单的 ZIP 压缩包
    2. **上传库存表** — HK库存日报（需含「⑤指定仓库分布」sheet）
    3. **点击执行** — 自动匹配 SKU，计算 TR 总库存 + 可借货仓满足情况
    4. **下载结果** — ZIP 包内含所有带 `_满足率` 后缀的订单文件
    
    ### 输出列说明
    | 列名 | 说明 |
    |------|------|
    | 匹配产品CODE | 库存表中匹配到的产品编码 |
    | 满足数量 | TR总库存可满足的数量 |
    | 来源仓别 | 满足数量的来源仓库 |
    | 需借货数量 | 仍需从借货仓补充的数量 |
    | 借货仓别 | 实际调用的借货仓库 |
    | 借货仓库存 | 借货仓的可借库存数量 |
    | REMARK | 其他可借货仓信息 |
    """)
