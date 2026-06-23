# pages/5_报表汇总.py - 显示最接近匹配
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO
import re
import difflib

# ---- [品牌映射、门店映射、辅助函数等与上一版本相同] ----
BRAND_NORMALIZE = {
    '4711': '4711', '4711香水': '4711',
    'ANNA SUI': 'ANNA SUI', '安娜苏': 'ANNA SUI', '安娜苏(香化）': 'ANNA SUI',
    '安娜苏（香化）': 'ANNA SUI', '安娜苏(香化)': 'ANNA SUI',
    'ANNA SUI 安娜苏': 'ANNA SUI', '安娜苏 香水': 'ANNA SUI',
    'VERSACE': 'VERSACE', '范思哲': 'VERSACE', 'VERSACE 范思哲': 'VERSACE',
    '范思哲 香水': 'VERSACE', '范思哲（香化）': 'VERSACE', '范思哲(香化）': 'VERSACE',
    'ATKINSONS': 'ATKINSONS', '阿特金森': 'ATKINSONS',
    'BABOR': 'BABOR',
    'Laura Mercier': 'Laura Mercier', '罗拉玛希': 'Laura Mercier',
    'MCM': 'MCM', '恩思恩': 'MCM', 'MCM 恩思恩': 'MCM',
    '恩思恩（香化）': 'MCM', '恩思恩(香化）': 'MCM',
    'PARFUMS de MARLY': 'PARFUMS de MARLY', '玛丽之香': 'PARFUMS de MARLY',
    'MOSCHINO': 'MOSCHINO', '默斯奇诺': 'MOSCHINO', 'MOSCHINO 默斯奇诺': 'MOSCHINO',
    '默斯奇诺（香化）': 'MOSCHINO', '默斯奇诺(香化）': 'MOSCHINO',
    'SALVATORE FERRAGAMO': 'SALVATORE FERRAGAMO', '菲拉格慕': 'SALVATORE FERRAGAMO',
    'Ferragamo': 'SALVATORE FERRAGAMO', 'FERRAGAMO': 'SALVATORE FERRAGAMO',
    'Ferragamo 菲拉格慕': 'SALVATORE FERRAGAMO', '菲拉格慕（香化）': 'SALVATORE FERRAGAMO',
    '菲拉格慕(香化）': 'SALVATORE FERRAGAMO',
    'ACCA KAPPA': 'ACCA KAPPA',
    'Clean': 'Clean', 'CLEAN': 'Clean', '克霖': 'Clean', 'CLEAN 克霖': 'Clean',
    'CLEAN 克霖 香水': 'Clean',
    'SANTA MONICA': 'SANTA MONICA', '圣曼尼加': 'SANTA MONICA',
    'LALIQUE': 'LALIQUE', '莱俪': 'LALIQUE', 'LALIQUE 莱俪': 'LALIQUE',
    '莱俪（香化）': 'LALIQUE', '莱俪(香化）': 'LALIQUE', 'LALIQUE 莱俪 香水': 'LALIQUE',
    'CHOPARD': 'CHOPARD', '萧邦': 'CHOPARD', 'Chopard 萧邦香水': 'CHOPARD',
    '萧邦（香化）': 'CHOPARD', '萧邦(香化）': 'CHOPARD',
    'DR.VRANJES': 'DR.VRANJES', 'Furla': 'Furla',
    'MICHAEL KORS': 'MICHAEL KORS', 'MICHAEL KORS(香化)': 'MICHAEL KORS',
    '[0442]MCM': 'MCM', '[0536]安娜苏': 'ANNA SUI', '[1321]罗拉玛希': 'Laura Mercier',
    '[1158]莱俪': 'LALIQUE', '[1157]4711': '4711', '[0459]范思哲': 'VERSACE',
    '[0485]默斯奇诺': 'MOSCHINO', '[1281]阿特金森': 'ATKINSONS', '[0559]菲拉格慕': 'SALVATORE FERRAGAMO',
}

STORE_MAP = {
    'CDF Hainan Membership Site': '补购', 'HTB': 'HTB', '前海 MPP': '前海 MPP',
    '7073 CHINA DUTY FREE(LANKA)(PVT)LTD': 'Sri Lanka',
    'Sanya AP International Departure': 'Sanya AP International Departure',
    'Haikou T2 International Departure': 'Haikou T2 International Departure',
    'XHG': 'XHG', 'Haikou Downtown': 'Haikou Downtown', 'Sanya AP': 'Sanya AP',
    'BoAo Downtown': 'BoAo Downtown', 'Erlian': 'Erlian', 'GLP': 'GLP',
    'Haikou AP': 'Haikou AP', 'Nanjing Arrival': 'Nanjing Arrival',
    'Mediterranea Cruise': 'Mediterranea Cruise', 'Piano Land cruise': 'Piano Land cruise',
    'Adora Cruise': 'Adora Cruise', 'SY DT': 'SYDT', 'MOVA DT': 'MOVA DT',
}

def get_normalized(brand):
    brand = brand.strip()
    if brand in BRAND_NORMALIZE: return BRAND_NORMALIZE[brand]
    for k, v in BRAND_NORMALIZE.items():
        if brand in k or k in brand: return v
    return brand

def parse_float(val):
    if val is None: return 0.0
    if isinstance(val, (int, float)): return float(val)
    if isinstance(val, str):
        val = val.strip().replace(',', '').replace(' ', '').replace('$', '').replace('¥', '')
        if val in ('', '-', '--'): return 0.0
        try: return float(val)
        except: return 0.0
    return 0.0

def extract_brand(filename):
    name = filename.replace('.xlsx', '').replace('.xls', '')
    patterns = [r'\s*[Ss]ales\s*[Rr]eport\s*\d{4}\.\d{2}\s*', r'\s*[Ss]ales\s*[Rr]eport\s*$',
                r'\s*销售报告\s*$', r'\s*\.\s*$']
    for p in patterns:
        name = re.sub(p, '', name)
    name = name.strip().strip('·').strip('-').strip()
    return name

def map_store(name):
    name = name.strip()
    if name in STORE_MAP: return STORE_MAP[name]
    if 'cruise' in name.lower(): return name
    return name

def clean_text(s):
    s = str(s)
    s = s.replace('（', '(').replace('）', ')')
    s = s.replace('(香化）', '(香化)').replace('（香化）', '(香化)')
    s = re.sub(r'\s+', ' ', s).strip()
    return s

# ---- CDFG ----
def parse_cdfg(file, month, year, cny_rate, hkd_rate):
    try:
        wb = load_workbook(file, data_only=True)
        ws = wb.active
    except: return []
    brand = extract_brand(file.name)
    col_a = []
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 5000), min_col=1, max_col=1, values_only=True):
        col_a.append(row[0])
    si = -1
    for i, v in enumerate(col_a):
        if v is not None and str(v).strip() == 'Summary': si = i; break
    if si < 0: return []
    recs = []
    i = si + 6
    seen = set()
    while i + 2 < len(col_a):
        loc = col_a[i]
        if loc is None: i += 1; continue
        s = str(loc).strip()
        if s in ('Barcode', 'Description', '合计', ''): break
        try: float(s); i += 1; continue
        except ValueError: pass
        if s in seen: break  # 详情区开始
        seen.add(s)
        a = parse_float(col_a[i+2])
        if a > 0: recs.append(('CDFG', map_store(s), brand, brand, year, month, a))
        i += 3
        if len(recs) > 80: break
    return recs

# ---- 颖通解析器（与之前相同）----
def parse_yt_haerbin(file, month, year, cny_rate, hkd_rate):
    try: wb = load_workbook(file, data_only=True); ws = wb.active
    except: return []
    tgt = None
    for row in ws.iter_rows(min_row=1, max_row=5, max_col=20, values_only=False):
        for cell in row:
            if cell.value and '3月销售数量' in str(cell.value): tgt = cell.column; break
        if tgt: break
    if not tgt: return []
    ac = tgt + 1; bt = {}
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=ac+2, values_only=True):
        if not row or not row[0]: continue
        f = str(row[0]).strip()
        if f in ('品牌名称','求和项:销售金额','','sku_id'): continue
        q = parse_float(row[tgt-1]) if tgt <= len(row) else 0
        a = parse_float(row[ac-1]) if ac <= len(row) else 0
        if q > 0 and a > 0:
            nb = get_normalized(f); bt[nb] = bt.get(nb,0) + a
    return [('WFJ','Harbin AP',b,b,year,month,a) for b,a in bt.items() if a>0]

def parse_yt_beijing(file, month, year, cny_rate, hkd_rate):
    try: df = pd.read_excel(file, header=None)
    except: return []
    uc = None
    for c in range(min(df.shape[1],20)):
        for r in range(min(10,df.shape[0])):
            if pd.notna(df.iloc[r,c]) and str(df.iloc[r,c]).strip()=='USD': uc=c; break
        if uc is not None: break
    if uc is None: return []
    recs = []
    for r in range(3, df.shape[0]):
        f = str(df.iloc[r,0]).strip() if pd.notna(df.iloc[r,0]) else ''
        if not f or f in ('品牌名称','求和项:库存数量3','求和项:销售数量3','求和项:零售金额3','USD','颖通',''): continue
        if f.startswith('0') or len(f)<2: continue
        usd = parse_float(df.iloc[r,uc]) if uc<df.shape[1] else 0
        if usd>0: nb = get_normalized(f); recs.append(('WFJ','BJ AP',f,nb,year,month,usd))
    return recs

def parse_yt_kuajing(file, month, year, cny_rate, hkd_rate):
    try: df = pd.read_excel(file, header=None)
    except: return []
    uc = None
    for c in range(min(df.shape[1],20)):
        for r in range(min(10,df.shape[0])):
            if pd.notna(df.iloc[r,c]) and str(df.iloc[r,c]).strip()=='USD': uc=c; break
        if uc is not None: break
    if uc is None: return []
    recs = []
    for r in range(5, df.shape[0]):
        f = str(df.iloc[r,0]).strip() if pd.notna(df.iloc[r,0]) else ''
        if not f or f in ('品牌名称','求和项:销量','求和项:销售额','USD',''): continue
        if f.startswith('0') or len(f)<2: continue
        usd = parse_float(df.iloc[r,uc]) if uc<df.shape[1] else 0
        if usd>0: nb = get_normalized(f); recs.append(('WFJ','Wanning DT',f,nb,year,month,usd))
    return recs

def parse_yt_manning(file, month, year, cny_rate, hkd_rate):
    try: import xlrd; wb = xlrd.open_workbook(file_contents=file.read()); ws = wb.sheet_by_index(0)
    except: return []
    bt = {}
    for r in range(1, ws.nrows):
        if ws.ncols<6: continue
        b = str(ws.cell_value(r,0)).strip()
        a = parse_float(ws.cell_value(r,5))
        if b and a>0: nb=get_normalized(b); bt[(b,nb)]=bt.get((b,nb),0)+a*cny_rate
    return [('颖通','万宁',b,nb,year,month,a) for (b,nb),a in bt.items() if a>0]

def parse_yt_normal(file, month, year, cny_rate, hkd_rate):
    try: import xlrd; wb = xlrd.open_workbook(file_contents=file.read()); ws = wb.sheet_by_index(0)
    except: return []
    bt = {}
    for r in range(1, ws.nrows):
        if ws.ncols<6: continue
        b = str(ws.cell_value(r,0)).strip()
        a = parse_float(ws.cell_value(r,5))
        if b and a>0: nb=get_normalized(b); bt[(b,nb)]=bt.get((b,nb),0)+a*cny_rate
    return [('颖通','',b,nb,year,month,a) for (b,nb),a in bt.items() if a>0]

def parse_yt_mar(file, month, year, cny_rate, hkd_rate):
    try: wb = load_workbook(file, data_only=True); ws = wb.active
    except: return []
    bt = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=10, values_only=True):
        if not row or not row[0]: continue
        b = str(row[0]).strip()
        if b in ('品牌','品牌名称',''): continue
        for c in range(3, min(len(row),10)):
            v = parse_float(row[c])
            if v>0: nb=get_normalized(b); bt[(b,nb)]=bt.get((b,nb),0)+v*cny_rate; break
    return [('颖通','',b,nb,year,month,a) for (b,nb),a in bt.items() if a>0]

# ---- Eternal ----
def parse_eternal(file, month, year, cny_rate, hkd_rate):
    try: wb = load_workbook(file, data_only=True); ws = wb.active
    except: return []
    hr, ac = None, None
    for row in ws.iter_rows(min_row=1, max_row=20, max_col=10, values_only=False):
        for cell in row:
            if cell.value and '月度销售金额' in str(cell.value): hr, ac = cell.row, cell.column; break
        if hr: break
    if not hr or not ac: return []
    bt = {}
    for row in ws.iter_rows(min_row=hr+1, max_row=ws.max_row, max_col=ac+2, values_only=True):
        if not row or not row[0]: continue
        b = str(row[0]).strip()
        if b in ('行标签','求和项:月度销售数量','求和项:月度销售金额(HKD)','求和项:期末数量','求和项:可周转月数','总计',''): continue
        hkd = parse_float(row[ac-1]) if ac<=len(row) else 0
        if hkd>0: nb=get_normalized(b); bt[nb]=bt.get(nb,0)+hkd*hkd_rate
    return [('ETERNAL','',b,b,year,month,a) for b,a in bt.items() if a>0]

def parse_eternal_gdf(file, month, year, cny_rate, hkd_rate):
    try: df = pd.read_excel(file, header=None)
    except: return []
    bc, ac = None, None
    for c in range(min(df.shape[1],30)):
        for r in range(min(5,df.shape[0])):
            v=str(df.iloc[r,c]).strip() if pd.notna(df.iloc[r,c]) else ''
            if v=='Brand': bc=c
            if v=='AMT-TTL': ac=c
    if bc is None or ac is None: return []
    bt = {}
    for r in range(5, df.shape[0]):
        b = str(df.iloc[r,bc]).strip() if pd.notna(df.iloc[r,bc]) else ''
        if not b or b in ('Brand','ETERNAL') or b.startswith('0') or len(b)<2: continue
        a = parse_float(df.iloc[r,ac]) if ac<df.shape[1] else 0
        if a>0: nb=get_normalized(b); bt[nb]=bt.get(nb,0)+a
    return [('ETERNAL','',b,b,year,month,a*hkd_rate) for b,a in bt.items() if a>0]

# ---- CNSC ----
def parse_cnsc(file, month, year, cny_rate, hkd_rate):
    try: df = pd.read_excel(file, header=None)
    except: return []
    bt = {}
    for r in range(3, df.shape[0]):
        n = str(df.iloc[r,0]).strip() if pd.notna(df.iloc[r,0]) else ''
        if not n or n in ('(空白)','门店','品牌','求和项:库存数量','求和项:总销售数量','求和项:总销售金额',
                           '求和项:线下销售数量','求和项:线下销售金额','求和项:电商销售数量','求和项:电商销售金额',''): continue
        if any(x in n for x in ('DT','AP','CQ')): continue
        usd = parse_float(df.iloc[r,9]) if df.shape[1]>9 else 0
        if usd>0: nb=get_normalized(n); bt[nb]=bt.get(nb,0)+usd
    return [('CNSC','SYDT',b,b,year,month,a) for b,a in bt.items() if a>0]

# ---- 文件路由器 ----
def get_parser(filename):
    name = filename.lower()
    if '美帕' in name or 'medspa' in name: return None, "SKIP"
    if '颖通' in name:
        if '哈尔滨' in name: return parse_yt_haerbin, None
        if '北京机场' in name: return parse_yt_beijing, None
        if '万宁' in name and name.endswith('.xls'): return parse_yt_manning, None
        if name.endswith('.xls'): return parse_yt_normal, None
        if '跨境' in name: return parse_yt_kuajing, None
        if '3月销售数据' in name: return parse_yt_mar, None
        return parse_yt_kuajing, None
    if 'eternal' in name:
        if '2026年3月销售报告' in name: return parse_eternal, None
        if 'gdf' in name: return parse_eternal_gdf, None
        return parse_eternal_gdf, None
    if 'cnsc' in name: return parse_cnsc, None
    if name.endswith('.xlsx'): return parse_cdfg, None
    return None, f"未能识别文件格式: {filename}"

# ============================================================
# 汇总表更新 - 带最近匹配建议
# ============================================================
def update_master(master_file, new_records, month, year):
    try:
        wb = load_workbook(master_file)
        if '2026 Full' in wb.sheetnames:
            ws = wb['2026 Full']
            st.success(f"✅ 使用Sheet: 2026 Full")
        else:
            ws = wb.active
            st.info(f"📋 使用Sheet: {ws.title}")
    except Exception as e:
        st.error(f"无法读取汇总表: {e}")
        return None
    
    # 搜索月份列
    target_names = [f'{month}月', ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'][month-1]]
    target_col = None
    header_row = None
    
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 100), max_col=min(ws.max_column+1, 50), values_only=False):
        for cell in row:
            if cell.value:
                val = str(cell.value).strip()
                if val in target_names:
                    target_col = cell.column
                    header_row = cell.row
                    break
        if target_col: break
    
    if target_col is None:
        st.error(f"找不到月份列。目标: {target_names}")
        for r in range(1, min(ws.max_row+1, 8)):
            vals = [str(ws.cell(row=r, column=c).value or '')[:15] for c in range(1, min(ws.max_column+1, 25))]
            st.write(f"行{r}: {vals}")
        return None
    
    st.success(f"✅ '{target_names[0]}' 在第{target_col}列")
    
    # 读取行标识
    row_data = []
    for r in range(header_row + 1, ws.max_row + 1):
        id_val = str(ws.cell(row=r, column=1).value or '').strip()
        if id_val:
            row_data.append((r, id_val))
    
    st.info(f"汇总表共 {len(row_data)} 行")
    
    # 显示前30行和包含"ETERNAL"/"颖通"/"安奈"等的行
    with st.expander("📋 查看汇总表行标识"):
        for rn, idv in row_data:
            st.write(f"行{rn}: {idv[:60]}")
    
    # 去重
    dedup = {}
    detail = []
    for rec in new_records:
        if len(rec) == 7:
            customer, store, orig_brand, norm_brand, y, m, amount = rec
        else:
            customer, store, norm_brand, y, m, amount = rec
            orig_brand = norm_brand
        key = (customer, store, norm_brand, y, m)
        if key in dedup: dedup[key] += amount
        else:
            dedup[key] = amount
            detail.append((customer, store, orig_brand, norm_brand, y, m))
    
    st.write(f"去重后: {len(dedup)} 条")
    
    # 匹配 + 显示最接近的行
    updated = 0
    not_found = []
    matched = []
    
    for (customer, store, orig_brand, norm_brand, y, m) in detail:
        key = (customer, store, norm_brand, y, m)
        amount = dedup[key]
        
        brands = list(set([b for b in [orig_brand, norm_brand] if b]))
        search_keys = set()
        
        for brand in brands:
            if store and store not in ('', 'SYDT'):
                search_keys.add(f"{customer}/{store}/{brand}")
                search_keys.add(f"{store}/{brand}")
                search_keys.add(brand)
            search_keys.add(f"{customer}/{brand}")
            search_keys.add(f"{customer} {brand}")
            search_keys.add(brand)
            # 去掉中文字符后的品牌名
            brand_en = re.sub(r'[\u4e00-\u9fff].*$', '', brand).strip()
            if brand_en and brand_en != brand:
                search_keys.add(brand_en)
        
        found = False
        used_search = ''
        used_row = ''
        
        for sk in search_keys:
            sk_l = sk.lower()
            for rn, id_val in row_data:
                id_l = id_val.lower()
                # 多种匹配方式
                if (sk_l in id_l or id_l in sk_l or 
                    sk_l.replace(' ','') in id_l.replace(' ','') or
                    id_l.replace(' ','') in sk_l.replace(' ','') or
                    sk_l.replace('/','') in id_l.replace('/','') or
                    id_l.replace('/','') in sk_l.replace('/','')):
                    ws.cell(row=rn, column=target_col, value=round(amount, 2))
                    updated += 1
                    used_search = sk
                    used_row = id_val[:40]
                    found = True
                    break
            if found: break
        
        if found:
            matched.append(f"✅ '{used_search}' → '{used_row}'")
        else:
            # 显示最接近的5行
            close_matches = []
            for sk in search_keys[:3]:  # 用前3个搜索词
                sk_l = sk.lower()
                for rn, id_val in row_data:
                    score = len(set(sk_l.split()) & set(id_val.lower().split()))
                    if score > 0:
                        close_matches.append((score, rn, id_val[:40]))
            close_matches = sorted(set(close_matches), key=lambda x: -x[0])[:5]
            
            not_found.append({
                'record': f"{customer}/{store}/{norm_brand} (${amount:,.2f})",
                'closest': close_matches
            })
    
    if matched:
        with st.expander(f"✅ 匹配 {len(matched)} 条"):
            for m in matched[:40]: st.write(m)
            if len(matched)>40: st.write(f"... 还有 {len(matched)-40} 条")
    
    if not_found:
        with st.expander(f"⚠️ {len(not_found)} 条未匹配（含最接近的行）"):
            for item in not_found:
                st.write(f"**{item['record']}**")
                if item['closest']:
                    st.write(f"  最接近的行:")
                    for score, rn, idv in item['closest'][:5]:
                        st.write(f"    行{rn}: '{idv}' (相似度:{score})")
                else:
                    st.write("  (无相似行)")
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output, updated, [item['record'] for item in not_found]

# ============================================================
# Session State & UI
# ============================================================
if 'processed' not in st.session_state:
    st.session_state.processed = False
    st.session_state.download_data = None
    st.session_state.download_name = ''
    st.session_state.preview_df = None
    st.session_state.updated = 0
    st.session_state.not_found = []

st.title("🔄 多零售商报表汇总工具")
st.markdown("---")

with st.sidebar:
    st.header("⚙️ 设置")
    month = st.selectbox("月份", range(1, 13), index=2)
    year = st.number_input("年份", min_value=2019, max_value=2030, value=2026)
    cny_rate = st.number_input("CNY→USD 汇率", value=0.1380, format="%.4f", step=0.001)
    hkd_rate = st.number_input("HKD→USD 汇率", value=0.1280, format="%.4f", step=0.001)

col1, col2 = st.columns([1, 1])

with col1:
    st.subheader("📋 步骤1：上传汇总表")
    master_f = st.file_uploader("选择现有汇总表", type=['xlsx', 'xls'], key="master")
    if master_f:
        st.success(f"✅ {master_f.name}")
        try:
            wb = load_workbook(master_f)
            st.info(f"📑 Sheets: {wb.sheetnames}")
        except: pass

with col2:
    st.subheader("📁 步骤2：上传零售商报表")
    sales_fs = st.file_uploader("选择零售商报表（可多选）", type=['xlsx', 'xls'],
                                 accept_multiple_files=True, key="sales")
    if sales_fs:
        st.info(f"已选 {len(sales_fs)} 个文件")

if master_f and sales_fs:
    st.markdown("---")
    btn = st.button("🚀 开始处理并更新汇总表", use_container_width=True, type="primary")
    
    if btn:
        all_records = []
        errors = []
        skipped = []
        
        bar = st.progress(0)
        status = st.empty()
        
        for i, f in enumerate(sales_fs):
            status.text(f"⏳ ({i+1}/{len(sales_fs)}) {f.name}")
            parser, err = get_parser(f.name)
            if parser is None:
                if err and err.startswith("SKIP"): skipped.append(f.name)
                else: errors.append(err or f"未能识别: {f.name}")
                bar.progress((i+1)/len(sales_fs))
                continue
            try:
                recs = parser(f, month, year, cny_rate, hkd_rate)
                if recs:
                    all_records.extend(recs)
                    st.write(f"✅ {f.name}: {len(recs)}条, ${sum(r[-1] for r in recs):,.2f}")
                else:
                    st.write(f"⚠️ {f.name}: 无有效数据")
            except Exception as e:
                errors.append(f"{f.name}: {e}")
                st.write(f"❌ {f.name}: {e}")
            bar.progress((i+1)/len(sales_fs))
        
        status.text("")
        
        if all_records:
            preview_data = [(r[0],r[1],r[2],r[4],r[5],r[6]) if len(r)==7 else r for r in all_records]
            st.success(f"✅ 共 {len(all_records)} 条记录")
            df = pd.DataFrame(preview_data, columns=['Customer','Stores','品牌','年份','月份','金额(USD)'])
            st.dataframe(df, use_container_width=True, height=300)
            
            st.info("正在更新汇总表...")
            result = update_master(master_f, all_records, month, year)
            
            if result:
                out, updated, not_found = result
                st.session_state.processed = True
                st.session_state.download_data = out.getvalue()
                st.session_state.download_name = f"01零售报表-品牌合计总表_{month}.{year}.xlsx"
                st.session_state.preview_df = df
                st.session_state.updated = updated
                st.session_state.not_found = not_found
                st.rerun()
        else:
            st.warning("没有有效记录")
        
        if errors:
            with st.expander(f"❌ 失败 {len(errors)} 个"):
                for e in errors: st.write(f"- {e}")
        if skipped:
            with st.expander(f"⏭️ 跳过 {len(skipped)} 个"):
                for s in skipped: st.write(f"- {s}")

if st.session_state.processed:
    st.markdown("---")
    st.subheader("✅ 处理完成")
    
    c1, c2 = st.columns(2)
    c1.metric("更新行数", st.session_state.updated)
    c2.metric("未匹配", len(st.session_state.not_found))
    
    if st.session_state.not_found:
        st.warning(f"⚠️ {len(st.session_state.not_found)} 条未匹配。展开上方'未匹配'区域查看最接近的行")
    
    if st.session_state.preview_df is not None:
        st.dataframe(st.session_state.preview_df, use_container_width=True, height=300)
    
    st.download_button("📥 下载更新后的汇总表",
        data=st.session_state.download_data,
        file_name=st.session_state.download_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True)
    
    if st.button("🔄 重新开始", use_container_width=True):
        for k in ['processed','download_data','download_name','preview_df','updated','not_found']:
            if k in st.session_state: del st.session_state[k]
        st.rerun()
